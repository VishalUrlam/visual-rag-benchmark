"""Export GQA benchmark results to Excel with embedded images."""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from .gqa_evaluator import GQAResult

_IMG_W = 120
_IMG_H = 90
_ROW_H = 70

_ANTHROPIC_MODEL = "Claude Opus 4.7"
_OPENAI_MODEL    = "GPT-5.5"


def _thumb(image_path: str, width: int, height: int) -> io.BytesIO:
    img = PILImage.open(image_path).convert("RGB")
    img.thumbnail((width * 2, height * 2), PILImage.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    buf.seek(0)
    return buf


def export_gqa_excel(
    anthropic_result: GQAResult | None,
    openai_result: GQAResult | None,
    output_path: Path,
    anthropic_model: str = _ANTHROPIC_MODEL,
    openai_model: str = _OPENAI_MODEL,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "GQA Results"

    dark_fill   = PatternFill("solid", fgColor="1F3864")
    anth_fill   = PatternFill("solid", fgColor="2E4057")
    oai_fill    = PatternFill("solid", fgColor="1B4332")
    white_bold  = Font(bold=True, color="FFFFFF", size=12)
    header_font = Font(bold=True, color="FFFFFF", size=11)

    # ── Row 1: model name banner ──────────────────────────────────────────── #
    ws.row_dimensions[1].height = 28

    for col in range(1, 6):
        c = ws.cell(row=1, column=col)
        c.fill = dark_fill

    if anthropic_result:
        ws.merge_cells("F1:G1")
        c = ws.cell(row=1, column=6, value=anthropic_model)
        c.fill = anth_fill
        c.font = white_bold
        c.alignment = Alignment(horizontal="center", vertical="center")

    if openai_result:
        ws.merge_cells("H1:I1")
        c = ws.cell(row=1, column=8, value=openai_model)
        c.fill = oai_fill
        c.font = white_bold
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 2: column headers ─────────────────────────────────────────────── #
    ws.row_dimensions[2].height = 30
    headers = [
        "Image", "Image ID", "Question", "Category", "Ground Truth",
        "Prediction", "✓ / ✗",
        "Prediction", "✓ / ✗",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = dark_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── column widths ─────────────────────────────────────────────────────── #
    col_widths = [18, 12, 40, 10, 16, 22, 8, 22, 8]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── data rows ─────────────────────────────────────────────────────────── #
    anth_map = {item.id: item for item in (anthropic_result.items if anthropic_result else [])}
    oai_map  = {item.id: item for item in (openai_result.items  if openai_result  else [])}
    all_ids  = list(dict.fromkeys(list(anth_map) + list(oai_map)))

    green = PatternFill("solid", fgColor="C6EFCE")
    red   = PatternFill("solid", fgColor="FFC7CE")

    for row_idx, qid in enumerate(all_ids, 3):
        anth = anth_map.get(qid)
        oai  = oai_map.get(qid)
        ref  = anth or oai

        ws.row_dimensions[row_idx].height = _ROW_H

        img_path = ref.image_path if ref else None
        if img_path and Path(img_path).exists():
            try:
                buf = _thumb(img_path, _IMG_W, _IMG_H)
                xl_img = XLImage(buf)
                xl_img.width  = _IMG_W
                xl_img.height = _IMG_H
                ws.add_image(xl_img, f"A{row_idx}")
            except Exception:
                ws.cell(row=row_idx, column=1, value="[image error]")
        else:
            ws.cell(row=row_idx, column=1, value="[missing]")

        for col, val in enumerate([
            None,
            ref.image_id if ref else "",
            ref.question if ref else "",
            ref.category if ref else "",
            ref.ground_truth if ref else "",
        ], 1):
            if val is not None:
                ws.cell(row=row_idx, column=col, value=val).alignment = Alignment(
                    vertical="center", wrap_text=True
                )

        def _write_pred(col: int, item) -> None:
            if item is None:
                ws.cell(row=row_idx, column=col, value="N/A").alignment = Alignment(vertical="center")
                ws.cell(row=row_idx, column=col + 1, value="").alignment = Alignment(vertical="center", horizontal="center")
                return
            ws.cell(row=row_idx, column=col, value=item.prediction).alignment = Alignment(
                vertical="center", wrap_text=True
            )
            tick = ws.cell(row=row_idx, column=col + 1, value="✓" if item.correct else "✗")
            tick.alignment = Alignment(horizontal="center", vertical="center")
            tick.fill = green if item.correct else red
            tick.font = Font(bold=True, color="375623" if item.correct else "9C0006")

        _write_pred(6, anth)
        _write_pred(8, oai)

    # ── summary ───────────────────────────────────────────────────────────── #
    summary_row = len(all_ids) + 4
    ws.cell(row=summary_row, column=1, value="Summary").font = Font(bold=True, size=12)

    if anthropic_result:
        ws.cell(row=summary_row + 1, column=1, value=anthropic_model)
        ws.cell(row=summary_row + 1, column=2, value=f"{anthropic_result.accuracy:.1%}")
    if openai_result:
        ws.cell(row=summary_row + 2, column=1, value=openai_model)
        ws.cell(row=summary_row + 2, column=2, value=f"{openai_result.accuracy:.1%}")

    wb.save(output_path)
