"""Export SEED-Bench MCQ results to Excel with embedded images."""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from .seedbench_evaluator import SEEDBenchResult

_IMG_W = 120
_IMG_H = 90
_ROW_H = 70

_MODEL_COLORS = {
    "anthropic":    "2E4057",
    "openai":       "1B4332",
    "gemini_pro":   "4A235A",
    "gemini_flash": "0E4D4D",
    "kimi":         "7B3F00",
}


def _thumb(image_path: str, width: int, height: int) -> io.BytesIO:
    img = PILImage.open(image_path).convert("RGB")
    img.thumbnail((width * 2, height * 2), PILImage.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    buf.seek(0)
    return buf


def export_seedbench_excel(
    output_path: Path,
    *,
    anthropic_result: SEEDBenchResult | None = None,
    openai_result: SEEDBenchResult | None = None,
    gemini_pro_result: SEEDBenchResult | None = None,
    gemini_flash_result: SEEDBenchResult | None = None,
    kimi_result: SEEDBenchResult | None = None,
    anthropic_model: str = "Claude Opus 4.7",
    openai_model: str = "GPT-5.5",
    gemini_pro_model: str = "Gemini 3.1 Pro",
    gemini_flash_model: str = "Gemini 3.5 Flash",
    kimi_model: str = "Kimi K2.6",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SEED-Bench Results"

    dark_fill  = PatternFill("solid", fgColor="1F3864")
    white_bold = Font(bold=True, color="FFFFFF", size=12)
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)

    # base cols: Image, Category, Level, Question, A, B, C, D = 8 cols
    BASE_COLS = 8

    models = []
    col = BASE_COLS + 1
    for label, result, color_key in [
        (anthropic_model,    anthropic_result,    "anthropic"),
        (openai_model,       openai_result,       "openai"),
        (gemini_pro_model,   gemini_pro_result,   "gemini_pro"),
        (gemini_flash_model, gemini_flash_result, "gemini_flash"),
        (kimi_model,         kimi_result,         "kimi"),
    ]:
        if result is not None:
            models.append((label, result, _MODEL_COLORS[color_key], col))
            col += 2

    # ── Row 1: banners ────────────────────────────────────────────────────── #
    ws.row_dimensions[1].height = 28
    for c in range(1, BASE_COLS + 1):
        ws.cell(row=1, column=c).fill = dark_fill

    for label, _, color, start_col in models:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 1)
        c = ws.cell(row=1, column=start_col, value=label)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = white_bold
        c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Row 2: headers ────────────────────────────────────────────────────── #
    ws.row_dimensions[2].height = 30
    base_headers = ["Image", "Category", "Level", "Question", "A", "B", "C", "D"]
    all_headers = base_headers + ["Prediction", "✓ / ✗"] * len(models)
    for c, h in enumerate(all_headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill = dark_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── column widths ─────────────────────────────────────────────────────── #
    base_widths = [18, 20, 6, 38, 22, 22, 22, 22]
    model_widths = [14, 8] * len(models)
    for c, w in enumerate(base_widths + model_widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── data rows ─────────────────────────────────────────────────────────── #
    all_maps = [({item.id: item for item in r.items}, start_col) for _, r, _, start_col in models]
    all_ids = list(dict.fromkeys(iid for m, _ in all_maps for iid in m))

    green    = PatternFill("solid", fgColor="C6EFCE")
    red      = PatternFill("solid", fgColor="FFC7CE")
    opt_fill = PatternFill("solid", fgColor="EBF3FB")

    for row_idx, qid in enumerate(all_ids, 3):
        ws.row_dimensions[row_idx].height = _ROW_H
        ref = next((m[qid] for m, _ in all_maps if qid in m), None)

        if ref and Path(ref.image_path).exists():
            try:
                buf = _thumb(ref.image_path, _IMG_W, _IMG_H)
                xl = XLImage(buf); xl.width, xl.height = _IMG_W, _IMG_H
                ws.add_image(xl, f"A{row_idx}")
            except Exception:
                ws.cell(row=row_idx, column=1, value="[error]")
        else:
            ws.cell(row=row_idx, column=1, value="[missing]")

        for c, val in enumerate([None,
            ref.category if ref else "",
            ref.level if ref else "",
            ref.question if ref else "",
        ], 1):
            if val is not None:
                ws.cell(row=row_idx, column=c, value=val).alignment = Alignment(vertical="center", wrap_text=True)

        if ref:
            gt = ref.ground_truth
            for c_idx, letter in enumerate("ABCD", 5):
                c = ws.cell(row=row_idx, column=c_idx, value=ref.options.get(letter, ""))
                c.alignment = Alignment(vertical="center", wrap_text=True)
                if letter == gt:
                    c.fill = PatternFill("solid", fgColor="D9EAD3")
                    c.font = Font(bold=True)
                else:
                    c.fill = opt_fill

        for item_map, start_col in all_maps:
            item = item_map.get(qid)
            if item is None:
                ws.cell(row=row_idx, column=start_col, value="N/A").alignment = Alignment(vertical="center")
                continue
            ws.cell(row=row_idx, column=start_col, value=item.prediction).alignment = Alignment(vertical="center", horizontal="center")
            tc = ws.cell(row=row_idx, column=start_col + 1, value="✓" if item.correct else "✗")
            tc.alignment = Alignment(horizontal="center", vertical="center")
            tc.fill = green if item.correct else red
            tc.font = Font(bold=True, color="375623" if item.correct else "9C0006")

    # ── summary ───────────────────────────────────────────────────────────── #
    sr = len(all_ids) + 4
    ws.cell(row=sr, column=1, value="Summary").font = Font(bold=True, size=12)
    for offset, (label, result, _, _) in enumerate(models, 1):
        ws.cell(row=sr + offset, column=1, value=label)
        ws.cell(row=sr + offset, column=2, value=f"{result.accuracy:.1%}")

    wb.save(output_path)
