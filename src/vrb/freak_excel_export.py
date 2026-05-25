"""Export FREAK benchmark results to Excel with embedded images (MCQ + QA sheets)."""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from .freak_evaluator import FREAKMCQResult, FREAKQAResult, MCQItemResult, QAItemResult
from .freak_loader import FREAKMCQSample, FREAKQASample

_IMG_W = 120
_IMG_H = 90
_ROW_H = 70

_MODEL_COLORS = {
    "anthropic":    "2E4057",
    "openai":       "1B4332",
    "gemini_pro":   "4A235A",
    "gemini_flash": "0E4D4D",
    "kimi":         "7B3F00",
}


def _thumb(image_path: Path, width: int, height: int) -> io.BytesIO:
    img = PILImage.open(image_path).convert("RGB")
    img.thumbnail((width * 2, height * 2), PILImage.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    buf.seek(0)
    return buf


def _build_model_list(
    anthropic_mcq, openai_mcq, gemini_pro_mcq, gemini_flash_mcq, kimi_mcq,
    anthropic_model, openai_model, gemini_pro_model, gemini_flash_model, kimi_model,
    base_cols: int,
) -> list[tuple[str, object, str, int]]:
    models = []
    col = base_cols + 1
    for label, result, color_key in [
        (anthropic_model,    anthropic_mcq,    "anthropic"),
        (openai_model,       openai_mcq,       "openai"),
        (gemini_pro_model,   gemini_pro_mcq,   "gemini_pro"),
        (gemini_flash_model, gemini_flash_mcq, "gemini_flash"),
        (kimi_model,         kimi_mcq,         "kimi"),
    ]:
        if result is not None:
            models.append((label, result, _MODEL_COLORS[color_key], col))
            col += 2
    return models


def _write_banner_and_headers(ws, models, base_headers, dark_fill, white_bold, hdr_font, base_cols):
    ws.row_dimensions[1].height = 28
    for c in range(1, base_cols + 1):
        ws.cell(row=1, column=c).fill = dark_fill

    for label, _, color, start_col in models:
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 1)
        c = ws.cell(row=1, column=start_col, value=label)
        c.fill = PatternFill("solid", fgColor=color)
        c.font = white_bold
        c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 30
    all_headers = base_headers + ["Prediction", "✓ / ✗"] * len(models)
    for c, h in enumerate(all_headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.fill = dark_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _write_mcq_sheet(
    ws,
    mcq_samples: list[FREAKMCQSample],
    models: list[tuple[str, FREAKMCQResult, str, int]],
) -> None:
    dark_fill  = PatternFill("solid", fgColor="1F3864")
    white_bold = Font(bold=True, color="FFFFFF", size=12)
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    green      = PatternFill("solid", fgColor="C6EFCE")
    red        = PatternFill("solid", fgColor="FFC7CE")
    opt_fill   = PatternFill("solid", fgColor="EBF3FB")

    BASE_COLS = 8  # Image, Category, Item, Question, A, B, C, D
    base_headers = ["Image", "Category", "Item", "Question", "A", "B", "C", "D"]
    _write_banner_and_headers(ws, models, base_headers, dark_fill, white_bold, hdr_font, BASE_COLS)

    base_widths = [18, 12, 14, 38, 20, 20, 20, 20]
    for c, w in enumerate(base_widths + [16, 8] * len(models), 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    sample_map = {s.id: s for s in mcq_samples}
    all_maps = [({item.id: item for item in r.items}, start_col) for _, r, _, start_col in models]
    all_ids = list(dict.fromkeys(iid for m, _ in all_maps for iid in m))

    for row_idx, sid in enumerate(all_ids, 3):
        ws.row_dimensions[row_idx].height = _ROW_H
        sample = sample_map.get(sid)

        if sample and sample.image_path.exists():
            try:
                buf = _thumb(sample.image_path, _IMG_W, _IMG_H)
                xl = XLImage(buf); xl.width, xl.height = _IMG_W, _IMG_H
                ws.add_image(xl, f"A{row_idx}")
            except Exception:
                ws.cell(row=row_idx, column=1, value="[error]")
        else:
            ws.cell(row=row_idx, column=1, value="[missing]")

        ref = next((m[sid] for m, _ in all_maps if sid in m), None)
        category = (sample.category if sample else "") or ""
        if isinstance(category, list):
            category = ", ".join(category)

        for c, val in enumerate([None, category,
            sample.item if sample else "",
            ref.question if ref else "",
        ], 1):
            if val is not None:
                ws.cell(row=row_idx, column=c, value=val).alignment = Alignment(vertical="center", wrap_text=True)

        if sample:
            for c_idx, opt in enumerate(sample.options[:4], 5):
                cell = ws.cell(row=row_idx, column=c_idx, value=opt)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if ref and opt == ref.ground_truth:
                    cell.fill = PatternFill("solid", fgColor="D9EAD3")
                    cell.font = Font(bold=True)
                else:
                    cell.fill = opt_fill

        for item_map, start_col in all_maps:
            item = item_map.get(sid)
            if item is None:
                ws.cell(row=row_idx, column=start_col, value="N/A").alignment = Alignment(vertical="center")
                continue
            ws.cell(row=row_idx, column=start_col, value=item.prediction).alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            tc = ws.cell(row=row_idx, column=start_col + 1, value="✓" if item.correct else "✗")
            tc.alignment = Alignment(horizontal="center", vertical="center")
            tc.fill = green if item.correct else red
            tc.font = Font(bold=True, color="375623" if item.correct else "9C0006")

    sr = len(all_ids) + 4
    ws.cell(row=sr, column=1, value="Summary").font = Font(bold=True, size=12)
    for offset, (label, result, _, _) in enumerate(models, 1):
        ws.cell(row=sr + offset, column=1, value=label)
        ws.cell(row=sr + offset, column=2, value=f"{result.accuracy:.1%}")


def _write_qa_sheet(
    ws,
    qa_samples: list[FREAKQASample],
    models: list[tuple[str, FREAKQAResult, str, int]],
) -> None:
    dark_fill  = PatternFill("solid", fgColor="1F3864")
    white_bold = Font(bold=True, color="FFFFFF", size=12)
    hdr_font   = Font(bold=True, color="FFFFFF", size=11)
    green      = PatternFill("solid", fgColor="C6EFCE")
    red        = PatternFill("solid", fgColor="FFC7CE")
    orange     = PatternFill("solid", fgColor="FFE0B2")

    BASE_COLS = 5  # Image, Category, Item, Question, Ground Truth
    base_headers = ["Image", "Category", "Item", "Question", "Ground Truth"]
    _write_banner_and_headers(ws, models, base_headers, dark_fill, white_bold, hdr_font, BASE_COLS)

    base_widths = [18, 12, 14, 38, 22]
    for c, w in enumerate(base_widths + [22, 10] * len(models), 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    sample_map = {s.id: s for s in qa_samples}
    all_maps = [({item.id: item for item in r.items}, start_col) for _, r, _, start_col in models]
    all_ids = list(dict.fromkeys(iid for m, _ in all_maps for iid in m))

    for row_idx, sid in enumerate(all_ids, 3):
        ws.row_dimensions[row_idx].height = _ROW_H
        sample = sample_map.get(sid)

        if sample and sample.image_path.exists():
            try:
                buf = _thumb(sample.image_path, _IMG_W, _IMG_H)
                xl = XLImage(buf); xl.width, xl.height = _IMG_W, _IMG_H
                ws.add_image(xl, f"A{row_idx}")
            except Exception:
                ws.cell(row=row_idx, column=1, value="[error]")
        else:
            ws.cell(row=row_idx, column=1, value="[missing]")

        ref = next((m[sid] for m, _ in all_maps if sid in m), None)
        category = (sample.category if sample else "") or ""
        if isinstance(category, list):
            category = ", ".join(category)

        for c, val in enumerate([None, category,
            sample.item if sample else "",
            ref.question if ref else "",
            sample.ground_truth if sample else "",
        ], 1):
            if val is not None:
                ws.cell(row=row_idx, column=c, value=val).alignment = Alignment(vertical="center", wrap_text=True)

        for item_map, start_col in all_maps:
            item = item_map.get(sid)
            if item is None:
                ws.cell(row=row_idx, column=start_col, value="N/A").alignment = Alignment(vertical="center")
                continue
            ws.cell(row=row_idx, column=start_col, value=item.prediction).alignment = Alignment(vertical="center", wrap_text=True)
            if item.correct:
                symbol, fill, color = "✓", green, "375623"
            elif item.hallucinated:
                symbol, fill, color = "HALLU", orange, "7B3F00"
            else:
                symbol, fill, color = "✗", red, "9C0006"
            tc = ws.cell(row=row_idx, column=start_col + 1, value=symbol)
            tc.alignment = Alignment(horizontal="center", vertical="center")
            tc.fill = fill
            tc.font = Font(bold=True, color=color)

    sr = len(all_ids) + 4
    ws.cell(row=sr, column=1, value="Summary").font = Font(bold=True, size=12)
    for offset, (label, result, _, _) in enumerate(models, 1):
        ws.cell(row=sr + offset, column=1, value=label)
        ws.cell(row=sr + offset, column=2, value=f"Acc {result.accuracy:.1%}  Hallu {result.hallucination_rate:.1%}")


def export_freak_excel(
    output_path: Path,
    mcq_samples: list[FREAKMCQSample],
    qa_samples: list[FREAKQASample],
    *,
    anthropic_mcq: FREAKMCQResult | None = None,
    openai_mcq: FREAKMCQResult | None = None,
    gemini_pro_mcq: FREAKMCQResult | None = None,
    gemini_flash_mcq: FREAKMCQResult | None = None,
    kimi_mcq: FREAKMCQResult | None = None,
    anthropic_qa: FREAKQAResult | None = None,
    openai_qa: FREAKQAResult | None = None,
    gemini_pro_qa: FREAKQAResult | None = None,
    gemini_flash_qa: FREAKQAResult | None = None,
    kimi_qa: FREAKQAResult | None = None,
    anthropic_model: str = "Claude Opus 4.7",
    openai_model: str = "GPT-5.5",
    gemini_pro_model: str = "Gemini 3.1 Pro",
    gemini_flash_model: str = "Gemini 3.5 Flash",
    kimi_model: str = "Kimi K2.6",
) -> None:
    wb = Workbook()

    # MCQ sheet
    ws_mcq = wb.active
    ws_mcq.title = "MCQ Results"
    mcq_models = _build_model_list(
        anthropic_mcq, openai_mcq, gemini_pro_mcq, gemini_flash_mcq, kimi_mcq,
        anthropic_model, openai_model, gemini_pro_model, gemini_flash_model, kimi_model,
        base_cols=8,
    )
    if mcq_models:
        _write_mcq_sheet(ws_mcq, mcq_samples, mcq_models)

    # QA sheet
    ws_qa = wb.create_sheet("QA Results")
    qa_models = _build_model_list(
        anthropic_qa, openai_qa, gemini_pro_qa, gemini_flash_qa, kimi_qa,
        anthropic_model, openai_model, gemini_pro_model, gemini_flash_model, kimi_model,
        base_cols=5,
    )
    if qa_models:
        _write_qa_sheet(ws_qa, qa_samples, qa_models)

    wb.save(output_path)
